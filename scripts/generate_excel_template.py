import os
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, ScatterChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart.label import DataLabelList

def create_template():
    wb = Workbook()
    
    # --- Sheet 1: DATA ---
    ws_data = wb.active
    ws_data.title = "DATA"
    # Placeholder headers and rows (these will be replaced by JS)
    headers = ["Category", "Value 1", "Value 2"]
    ws_data.append(headers)
    for i in range(1, 11):
        ws_data.append([f"Item {i}", i * 10, i * 15])
        
    for col in ws_data.columns:
        ws_data.column_dimensions[col[0].column_letter].width = 15

    # --- Sheet 2: CHART_DATA ---
    ws_chart_data = wb.create_sheet(title="CHART_DATA")
    # This sheet holds aggregated data for the charts.
    chart_headers = ["Label", "Metric"]
    ws_chart_data.append(chart_headers)
    for i in range(1, 11):
        ws_chart_data.append([f"Label {i}", i * 100])

    for col in ws_chart_data.columns:
        ws_chart_data.column_dimensions[col[0].column_letter].width = 15

    # --- Sheet 3: DASHBOARD ---
    ws_dashboard = wb.create_sheet(title="DASHBOARD")
    ws_dashboard.sheet_view.showGridLines = False

    # Styling for the dashboard background and title
    title_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    title_font = Font(name="Arial", size=18, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")

    ws_dashboard.merge_cells("B2:P3")
    title_cell = ws_dashboard["B2"]
    title_cell.value = "Interactive Data Dashboard"
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = align_center

    # Add a subtitle/metadata row
    ws_dashboard.merge_cells("B4:P4")
    sub_cell = ws_dashboard["B4"]
    sub_cell.value = "Data sourced from VisualizeMyData - Fully Editable Excel Charts"
    sub_cell.font = Font(name="Arial", size=11, italic=True, color="64748B")
    sub_cell.alignment = align_center

    # We will define 5 charts. The JS code will delete the ones not selected.
    # Max rows for charts to reference (up to 100 rows by default).
    max_row = 101

    cats = Reference(ws_chart_data, min_col=1, min_row=2, max_row=max_row)
    data = Reference(ws_chart_data, min_col=2, min_row=1, max_row=max_row)

    # 1. Bar Chart
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Bar Chart"
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.shape = 4
    bar_chart.width = 15
    bar_chart.height = 10
    ws_dashboard.add_chart(bar_chart, "B6")

    # 2. Line Chart
    line_chart = LineChart()
    line_chart.title = "Line Chart"
    line_chart.style = 13
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(cats)
    line_chart.width = 15
    line_chart.height = 10
    ws_dashboard.add_chart(line_chart, "I6")

    # 3. Pie Chart
    pie_chart = PieChart()
    pie_chart.title = "Pie Chart"
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(cats)
    pie_chart.width = 15
    pie_chart.height = 10
    # Enable data labels as percentages
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    ws_dashboard.add_chart(pie_chart, "B22")

    # 4. Area Chart
    area_chart = AreaChart()
    area_chart.title = "Area Chart"
    area_chart.style = 13
    area_chart.add_data(data, titles_from_data=True)
    area_chart.set_categories(cats)
    area_chart.width = 15
    area_chart.height = 10
    ws_dashboard.add_chart(area_chart, "I22")

    # 5. Scatter Chart (requires numeric X and Y, using DATA sheet)
    scatter_chart = ScatterChart()
    scatter_chart.title = "Scatter Chart"
    scatter_chart.style = 13
    xvalues = Reference(ws_data, min_col=2, min_row=2, max_row=max_row)
    yvalues = Reference(ws_data, min_col=3, min_row=2, max_row=max_row)
    series = Series(yvalues, xvalues, title_from_data=False)
    scatter_chart.series.append(series)
    scatter_chart.width = 15
    scatter_chart.height = 10
    ws_dashboard.add_chart(scatter_chart, "B38")

    # Ensure the templates dir exists
    out_dir = "public/templates"
    os.makedirs(out_dir, exist_ok=True)
    
    out_path = os.path.join(out_dir, "dashboard-template.xlsx")
    wb.save(out_path)
    print(f"Successfully generated {out_path}")

if __name__ == "__main__":
    create_template()
